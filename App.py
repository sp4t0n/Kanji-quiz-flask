from pywebio import start_server
from pywebio.input import select, input, checkbox, TEXT, SELECT
from pywebio.output import put_text, put_buttons, put_markdown, put_error, use_scope, put_html
from pywebio.session import hold
from openpyxl import Workbook, load_workbook
import os
import random

DATA_FILE = "quiz_data.xlsx"

class QuizApp:
    def __init__(self):
        self.initialize_variables()
        self.load_quiz_data()
        self.wrong_answers = []
        self.showing_errors = False

    def initialize_variables(self):
        self.quiz_categories = []
        self.quiz_data = {}
        self.current_category = ""
        self.current_quiz = None
        self.correct_answers = 0
        self.total_questions = 0
        self.shown_quizzes = {}
        self.quiz_direction = 'kanji to meaning'
        self.selected_categories = []
        self.last_output = None
        self.score_output = None
        self.show_romaji = False
        self.select_all_clicked = False

    def create_empty_data_file(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'
        headers = ['Kanji', 'Romanji', 'Significato', 'Categoria', 'Tipo (Verbo v /Aggettivo a)']
        for idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=header)
            ws.column_dimensions[chr(64 + idx)].width = 30
        wb.save(DATA_FILE)
        self.quiz_data['Generale'] = []

    def read_data_from_file(self):
        wb = load_workbook(DATA_FILE)
        ws = wb.active
        all_quizzes = []
        for row in ws.iter_rows(min_row=2):
            kanji, romaji, meaning, category, quiz_type = (
                row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
            if quiz_type:
                quiz_type = quiz_type.lower()
            if not category or category == "Categoria":
                category = "Generale"
            all_quizzes.append({'kanji': kanji, 'romaji': romaji,
                                'meaning': meaning, 'category': category, 'type': quiz_type})
        all_quizzes.sort(key=lambda x: x['category'])
        self.quiz_data = {}
        for quiz in all_quizzes:
            category = quiz['category']
            if category not in self.quiz_data:
                self.quiz_data[category] = []
            self.quiz_data[category].append(quiz)
        self.quiz_categories = list(self.quiz_data.keys())

    def load_quiz_data(self):
        try:
            if not os.path.isfile(DATA_FILE):
                self.create_empty_data_file()
            else:
                self.read_data_from_file()
        except Exception as e:
            put_error(f"Errore durante il caricamento dei dati del quiz: {str(e)}")

    def save_quiz_data(self):
        try:
            if os.path.isfile(DATA_FILE):
                wb = load_workbook(DATA_FILE)
            else:
                wb = Workbook()

            if "QuizData" in wb.sheetnames:
                ws = wb["QuizData"]
            else:
                ws = wb.create_sheet(title="QuizData")

            # Cancella tutte le righe esistenti prima di scrivere nuovi dati
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            # Scrivi l'intestazione delle colonne
            headers = ['Kanji', 'Romaji', 'Significato', 'Categoria', 'Tipo']
            for col_num, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_num, value=header)

            # Scrivi i nuovi dati
            row_num = 2
            for category, quizzes in self.quiz_data.items():
                for quiz in quizzes:
                    if not category:
                        category = "Generale"
                    for col_num, value in enumerate([quiz['kanji'], quiz['romaji'], quiz['meaning'], category, quiz['type']], start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                    row_num += 1

            wb.save(DATA_FILE)

        except PermissionError:
            put_error("Impossibile salvare i dati del quiz.")
        except Exception as e:
            put_error(f"Si è verificato un errore durante il salvataggio dei dati del quiz: {str(e)}")

    def next_random_quiz(self):
        if not hasattr(self, 'selected_categories') or not self.selected_categories:
            self.show_category_checkboxes()
            
        if not self.selected_categories:
            put_text("Devi selezionare almeno una categoria.")
            return None

        # Controlla se tutti i quiz di tutte le categorie selezionate sono stati utilizzati
        all_done = all(len(self.shown_quizzes.get(cat, set())) == len(self.quiz_data.get(cat, [])) for cat in self.selected_categories)

        if all_done:
            # Mostra un messaggio di completamento e chiedi all'utente di selezionare nuove categorie
            put_text("🎉🎉🎉QUIZ COMPLETATO🎉🎉🎉")
            self.shown_quizzes = {cat: set() for cat in self.selected_categories}  # Resetta i quiz mostrati per le categorie selezionate
            self.show_category_checkboxes()
            return None

        # Prendi una categoria a caso tra quelle selezionate che hanno ancora quiz disponibili
        available_categories = [cat for cat in self.selected_categories if len(self.shown_quizzes.get(cat, set())) < len(self.quiz_data.get(cat, []))]
        if not available_categories:
            put_text("Nessun quiz disponibile nelle categorie selezionate.")
            return None

        self.current_category = random.choice(available_categories)

        remaining_quizzes = [
            q for i, q in enumerate(self.quiz_data[self.current_category]) if i not in self.shown_quizzes.get(self.current_category, set())]

        quiz = random.choice(remaining_quizzes)
        quiz_index = self.quiz_data[self.current_category].index(quiz)
        self.shown_quizzes.setdefault(self.current_category, set()).add(quiz_index)
        return quiz

    def toggle_romaji(self, clicked_button_value=None):
        self.show_romaji = not self.show_romaji
        with use_scope('romaji', clear=True):
            if self.show_romaji:
                put_text(f"Romaji: {self.current_quiz['romaji']}")

    def next_question(self):
        with use_scope('question', clear=True):
            next_quiz = self.next_random_quiz()
            
            if next_quiz is None:
                return

            self.current_quiz = next_quiz
            self.display_question_based_on_direction()

            with use_scope('romaji', clear=True):
                if self.show_romaji:
                    put_text(f"Romaji: {self.current_quiz['romaji']}")

    def display_question_based_on_direction(self):
        def get_type(quiz):
            return f" ({quiz['type']})" if quiz['type'] else ""

        is_kanji_format = bool(self.current_quiz['kanji'])
        possible_wrong_answers = [quiz for quiz in self.quiz_data[self.current_category] if quiz !=
                                  self.current_quiz and quiz.get('type') == self.current_quiz.get('type') and bool(quiz['kanji']) == is_kanji_format]

        if len(possible_wrong_answers) < 2:
            all_other_quizzes = [quiz for cat, quizzes in self.quiz_data.items() for quiz in quizzes if cat !=
                                 self.current_category and quiz != self.current_quiz and quiz.get('type') == self.current_quiz.get('type') and bool(quiz['kanji']) == is_kanji_format]
            wrong_answers = random.sample(
                all_other_quizzes, 2 - len(possible_wrong_answers))
            wrong_answers.extend(possible_wrong_answers)
        else:
            wrong_answers = random.sample(possible_wrong_answers, 2)

        if self.quiz_direction == 'kanji to meaning':
            question_format = '<span style="color: red; font-size: 24px;">Quale è il significato di questo kanji/katakana: {}?</span>'
            question_content = self.current_quiz['kanji'] if self.current_quiz['kanji'] else self.current_quiz['romaji']
            options = [
                self.current_quiz['meaning'] + get_type(self.current_quiz),
                wrong_answers[0]['meaning'] + get_type(wrong_answers[0]),
                wrong_answers[1]['meaning'] + get_type(wrong_answers[1])
            ]
        else:
            question_format = '<span style="color: blue; font-size: 24px;">Quale kanji/katakana corrisponde a questo significato: {}?</span>'
            question_content = self.current_quiz['meaning']
            options = [
                (self.current_quiz['kanji'] if self.current_quiz['kanji'] else self.current_quiz['romaji']) + get_type(self.current_quiz),
                (wrong_answers[0]['kanji'] if wrong_answers[0]['kanji'] else wrong_answers[0]['romaji']) + get_type(wrong_answers[0]),
                (wrong_answers[1]['kanji'] if wrong_answers[1]['kanji'] else wrong_answers[1]['romaji']) + get_type(wrong_answers[1])
            ]

        put_html(question_format.format(question_content))
        random.shuffle(options)
        put_buttons(options, onclick=self.check_answer)

    def check_answer(self, selected_option):
        with use_scope('feedback', clear=True):
            correct_answer = self.get_correct_answer()
            question_text = self.get_question_text()
            romaji = self.get_romaji()  # Ottieni il romaji associato

            if selected_option == correct_answer:
                self.correct_answers += 1
                put_text("Risposta esatta! ✅")
            else:
                put_html(f"<div><span style='color: red;'>Risposta errata!</span> ❌<br><span style='color: blue;'>La domanda era:</span> '{question_text}'.<br><span style='color: green;'>La risposta corretta era:</span> {correct_answer}.</div>")
                # Memorizza la domanda, la risposta corretta, la risposta fornita e il romaji
                self.wrong_answers.append({
                    'question': question_text,
                    'correct_answer': correct_answer,
                    'given_answer': selected_option,
                    'romaji': romaji  # Memorizza il romaji
                })

        self.total_questions += 1
        self.update_score()
        self.next_question()

    def get_romaji(self):
        return self.current_quiz['romaji']

    def show_error_recap(self):
        with use_scope('errors', clear=True):  # Usa un nuovo scope per gli errori
            if not self.showing_errors:
                for error in self.wrong_answers:
                    if self.direction == "kanji_to_romaji":
                        question_with_romaji = f"{error['question']} ({error['romaji']})"
                        put_html(f"<span style='color: blue;'>Domanda:</span> {question_with_romaji}<br>")
                    else:
                        correct_answer_with_romaji = f"{error['correct_answer']} ({error['romaji']})"
                        put_html(f"<span style='color: blue;'>Domanda:</span> {error['question']}<br>")
                        put_html(f"<span style='color: green;'>Risposta corretta:</span> {correct_answer_with_romaji}<br>")
                    put_html(f"<span style='color: red;'>Risposta fornita:</span> {error['given_answer']}<br>")
                    put_markdown("---")
                self.showing_errors = True
            else:
                self.showing_errors = False
    
    
    def get_correct_answer(self):
        if self.quiz_direction == 'kanji to meaning':
            correct_answer = self.current_quiz['meaning']
        else:
            correct_answer = self.current_quiz['kanji'] if self.current_quiz['kanji'] else self.current_quiz['romaji']

        if self.current_quiz.get('type'):
            correct_answer += f" ({self.current_quiz['type']})"
        return correct_answer

    def get_question_text(self):
        if self.quiz_direction == 'kanji to meaning':
            question_text = f"Quale è il significato di questo kanji/katakana: {self.current_quiz['kanji'] if self.current_quiz['kanji'] else self.current_quiz['romaji']}?"
        else:
            question_text = f"Quale kanji/katakana corrisponde a questo significato: {self.current_quiz['meaning']}?"
        return question_text

    def update_score(self):
        with use_scope('score', clear=True):
            correct_percentage = (self.correct_answers / self.total_questions) * 100 if self.total_questions else 0
            wrong_percentage = 100 - correct_percentage
                        
            progress_bar = f"""
            <div style="background-color: #f3f3f3; border-radius: 10px; padding: 3px;" onclick="show_error_recap()">
                <div style="display: flex; align-items: center; justify-content: center;">
                    <div style="width: {correct_percentage}%; background-color: #4CAF50; text-align: center; padding: 10px 0; border-radius: 8px;">
                        {self.correct_answers}
                    </div>
                    <div style="width: {wrong_percentage}%; background-color: #FF5733; text-align: center; padding: 10px 0; border-radius: 8px;">
                        {self.total_questions - self.correct_answers}
                    </div>
                </div>
            </div>
            """
            put_html(progress_bar)


    def reset_score(self):
        self.correct_answers = 0
        self.total_questions = 0
        self.update_score()
        self.wrong_answers = []  # Aggiunto per eliminare gli errori salvati
        self.showing_errors = False  # Aggiunto per nascondere gli errori mostrati
        with use_scope('errors', clear=True):  # Aggiunto per pulire l'area degli errori
            pass

    def switch_mode(self):
        if self.quiz_direction == 'kanji to meaning':
            self.quiz_direction = 'meaning to kanji'
        else:
            self.quiz_direction = 'kanji to meaning'
        self.next_question()

    def clear_categories(self):
        if hasattr(self, 'selected_categories'):
            delattr(self, 'selected_categories')
        put_text("Le categorie selezionate sono state cancellate. Sarai in grado di selezionarne di nuove.")

    def handle_category_selection(self, selected_categories):
        if selected_categories:
            self.selected_categories = selected_categories
            self.current_category = random.choice(self.selected_categories)
            self.next_question()
        else:
            put_text("Devi selezionare almeno una categoria.")

    def select_all_categories(self):
        self.selected_categories = self.quiz_categories.copy()
        self.show_category_checkboxes()

    def show_category_checkboxes(self):
        options = ['Seleziona tutto'] + self.quiz_categories
        selected_categories = checkbox("Seleziona una o più categorie", options=options, value=self.selected_categories)
        
        if 'Seleziona tutto' in selected_categories:
            self.selected_categories = self.quiz_categories.copy()
            self.handle_category_selection(self.selected_categories)  # Simula un clic su "Submit"
            return

        self.handle_category_selection(selected_categories)

    def add_category(self):
        category = input("Aggiungi Categoria", type=TEXT, placeholder="Inserisci il nome della categoria")
        if category:
            if category not in self.quiz_categories:
                self.quiz_categories.append(category)
                self.quiz_data[category] = []
                put_text('La categoria è stata aggiunta con successo!')
                self.save_quiz_data()
            else:
                put_error('Errore: La categoria esiste già.')

    def edit_category(self):
        selected_category = select("Seleziona una categoria da modificare", type=SELECT, options=self.quiz_categories)
        if selected_category:
            new_category_name = input(f"Modifica il nome della categoria '{selected_category}':", type=TEXT)
            if new_category_name:
                if new_category_name not in self.quiz_categories:
                    self.quiz_data[new_category_name] = self.quiz_data.pop(selected_category)
                    self.quiz_categories.remove(selected_category)
                    self.quiz_categories.append(new_category_name)
                    self.quiz_categories.sort()
                    self.save_quiz_data()
                else:
                    put_error('Errore: La categoria esiste già.')

    def edit_quiz(self):
        selected_category = select("Seleziona una categoria per modificare un quiz", type=SELECT, options=self.quiz_categories)
        if selected_category:
            if selected_category not in self.quiz_data or not self.quiz_data[selected_category]:
                put_text('La categoria selezionata non contiene ancora quiz. Aggiungine uno prima di modificarlo.')
                return

            quiz_list = [f"{quiz['kanji']} - {quiz['meaning']}" for quiz in self.quiz_data[selected_category]]
            selected_quiz_str = select("Seleziona un quiz da modificare", type=SELECT, options=quiz_list)

            # Trova il quiz selezionato
            for quiz in self.quiz_data[selected_category]:
                if f"{quiz['kanji']} - {quiz['meaning']}" == selected_quiz_str:
                    selected_quiz = quiz
                    break

        if selected_quiz is None:
            put_text("Quiz non trovato.")
            return

        # Utilizzo di input singoli invece di input_group
        kanji = input("Modifica Kanji:", type=TEXT, value=selected_quiz.get('kanji', ''))
        meaning = input("Modifica Significato:", type=TEXT, value=selected_quiz.get('meaning', ''))
        romaji = input("Modifica Romaji:", type=TEXT, value=selected_quiz.get('romaji', ''))
        quiz_type = input("Modifica Tipo (a/v):", type=TEXT, value=selected_quiz.get('type', ''))

        # Aggiorna il quiz
        selected_quiz.update({'kanji': kanji, 'meaning': meaning, 'romaji': romaji, 'type': quiz_type})
        self.save_quiz_data()

    def add_quiz(self):
        selected_category = select("Seleziona una categoria per aggiungere un quiz", type=SELECT, options=self.quiz_categories)
        if selected_category:
            kanji = input("Inserisci Kanji:", type=TEXT)
            meaning = input("Inserisci Significato:", type=TEXT)
            romaji = input("Inserisci Romaji:", type=TEXT)
            quiz_type = input("Inserisci Tipo (a/v):", type=TEXT)
            
            if romaji and meaning:
                self.quiz_data[selected_category].append({'kanji': kanji, 'romaji': romaji, 'meaning': meaning, 'category': selected_category, 'type': quiz_type})
                self.save_quiz_data()
            else:
                put_error('Errore: Inserisci sia il kanji che il significato.')


def main():
    quiz_app = QuizApp()
    put_markdown("# Benvenuto a Kanji Quiz!")
    quiz_app.show_category_checkboxes()
    
    put_markdown("---")
    
    def reset_button_click():
        quiz_app.reset_score()

    with use_scope('score'):
        quiz_app.update_score()

    with use_scope('feedback', clear=True):
        pass

    put_buttons(['Resetta Punteggio/Errori', 'Mostra/Nascondi Romaji', 'Mostra/Nascondi Errori'], onclick=[quiz_app.reset_score, quiz_app.toggle_romaji, quiz_app.show_error_recap])  # Rinominato il pulsante in "Mostra/Nascondi Errori"
    put_markdown("---")
    put_text("🛠️ Impostazioni:")
    with use_scope('settings_buttons'):
        put_buttons(['Cambia modalità', 'Quiz Successivo', 'Cambia categorie'], onclick=[quiz_app.switch_mode, quiz_app.next_question, quiz_app.show_category_checkboxes])
                
    put_markdown("---")
    put_text("🗄️ Edit:")
    with use_scope('edit_buttons'):
        put_buttons(['Aggiungi Categoria', 'Modifica categoria', 'Aggiungi Quiz', 'Modifica Quiz'], onclick=[quiz_app.add_category, quiz_app.edit_category, quiz_app.add_quiz, quiz_app.edit_quiz])        
    
    put_markdown("---")
    
    def toggle_dark_mode(btn_val=None):
        # Usa JavaScript per aggiungere/rimuovere la classe "dark-mode" al body
        toggle_js = """
        if (document.body.classList.contains('dark-mode')) {
            document.body.classList.remove('dark-mode');
        } else {
            document.body.classList.add('dark-mode');
        }
        """
        put_html(f"<script>{toggle_js}</script>")


    dark_mode_css = """
    <style>
        /* Stili per la modalità scura */
        body.dark-mode {
            background-color: #121212;
            color: white;
        }
        
        /* Stile per input-container in modalità scura */
        body.dark-mode #input-container {
            background-color: #121212;  /* Scegli un colore di sfondo appropriato */
            /* Aggiungi altri stili se necessario */
        }

        /* Stile per .card e i suoi figli in modalità scura */
        body.dark-mode .card, 
        body.dark-mode .card * {
            background-color: #121212;  /* Scegli un colore di sfondo appropriato */
            color: white;  /* Imposta il colore del testo a bianco */
        }

        /* Aggiungi altri stili per la modalità scura qui */
    </style>
    """
    put_html(dark_mode_css)
    
    put_buttons(['Modalità Notte/Giorno'], onclick=toggle_dark_mode)    
    hold()

if __name__ == "__main__":
    start_server(main, host='0.0.0.0', debug=True, port=80)
